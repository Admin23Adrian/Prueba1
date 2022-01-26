import shutil
from zipfile import LargeZipFile
from autenticarse import loguearse
from descarga_remito import descarga_remitos
import openpyxl
import pythoncom
from getpass import getuser
import time
import os
from datetime import date
from datetime import datetime
import tkinter as tk
from tkinter import ttk
from tkinter import *
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

#-------------------------------------------------Interfaz----------------------------------------------------------#
#Espacio para interfaz


def pdf_cap():
#-------------------------------------------------Ventana Raiz------------------------------------------------------#
    user = getuser()
    root = Tk()
    dt = datetime.now()    # Fecha y hora actual
    day = dt.day
    month = dt.month
    year = dt.year
    hoy = str(day) + "-" + str(month) + "-" + str(year)

    token = loguearse()

    pythoncom.CoInitialize()
    path=('C:/Users/' + user +  '/Desktop/pdf_para_cap/liquidacion.xlsx')
    pathpdf='U:/Administracion_y_Finanzas/Cuentas_a_Pagar/27 Liquidacion Facturas de RdF/HISTORIAL_PDF/'
    pathexcel='U:/Administracion_y_Finanzas/Cuentas_a_Pagar/27 Liquidacion Facturas de RdF/HISTORIAL_EXCEL'
    
    wb = openpyxl.load_workbook(path,keep_vba=False)
    ws = wb["CaP"]
    cuits = []
    entregas = []
    fechas = []
    ultimafiladelws = len(ws['A'])
    fecha_excel = ""
    
    # ARMAR CARPETA PARA CADA LEGAJO.
    for dato in range(2, ultimafiladelws + 1):
        if dato == None:
            continue
        else:
            cuits = ws.cell(row = dato, column = 3).value
            entregas = str(ws.cell(row = dato,column = 1).value)
            fechas = ws.cell(row = dato,column = 7).value
            
            str_fecha = datetime.strftime(fechas, '%d-%m-%Y')
            fecha_excel = str_fecha
            
            pathcuits = 'U:/Administracion_y_Finanzas/Cuentas_a_Pagar/27 Liquidacion Facturas de RdF/HISTORIAL_PDF/' + str(cuits)            
            pathfechas = 'U:/Administracion_y_Finanzas/Cuentas_a_Pagar/27 Liquidacion Facturas de RdF/HISTORIAL_PDF/' + str(cuits)+ '/' + str_fecha
            
            print(f"\n ----- ENTREGA: {entregas} -----")
            if os.path.isdir(pathcuits):
                if not os.path.isdir(pathfechas):
                    print(f">> CREANDO CARPETA DE FECHAS.")
                    os.mkdir(pathfechas)
                    time.sleep(1)
                    descarga_remitos(entregas, token, pathfechas)
                else:
                    print(f"La carpeta {pathfechas} ya existe. Se procede a descargar la info aqui dentro.")
                    descarga_remitos(entregas, token, pathfechas)
            else:
                print(f"--->>> CREANDO CARPETA DE CUITS {pathcuits}.")
                os.mkdir(pathcuits)
                os.mkdir(pathfechas)
                time.sleep(1)
                descarga_remitos(entregas, token, pathfechas)
     
    wb.close()
    try:
        # ARMAR CARPETA PARA CADA EXCEL.
        pathfechas2 = f"{pathexcel}/{str_fecha}"            
        os.mkdir(pathfechas2)
        shutil.copy(path,pathfechas2)
    except:
        print(">>> Algo salio mal en la creacion de carpeta de Excel.")          
            
#---------------------------------------------------------fin-----------------------------------------------------------#
if __name__=="__main__":
    pdf_cap()    